"""
Preprocesses raw data into data/processed/ for the FPFA Carbon Intensity & Macrogrid Tool.
Run once before serving: python build.py
"""

import json
import sys
from pathlib import Path

import geopandas as gpd
import numpy as np
import openpyxl
import pandas as pd

RAW = Path("data/raw")
OUT = Path("data/processed")
OUT.mkdir(parents=True, exist_ok=True)

NAICS_SECTOR = {
    "3311": "Iron & Steel",
    "3312": "Iron & Steel",
    "3313": "Aluminum",
    "3314": "Aluminum",
    "32731": "Cement",
    "32721": "Glass",
    "32722": "Glass",
    "32531": "Fertilizer",
    "32512": "Hydrogen",
    "33441": "Solar",
}

def naics_sector(code):
    code = str(code).strip()
    for prefix, label in NAICS_SECTOR.items():
        if code.startswith(prefix):
            return label
    return None


# ── 1a. SHP → GeoJSON ────────────────────────────────────────────────────────

def build_geojson():
    shp = RAW / "egrid2023_subregions" / "eGRID2023_Subregions.shp"
    gdf = gpd.read_file(shp)
    if gdf.crs.to_epsg() != 4326:
        gdf = gdf.to_crs(epsg=4326)
    gdf["geometry"] = gdf["geometry"].simplify(tolerance=0.01, preserve_topology=True)
    out = OUT / "egrid_subregions.geojson"
    gdf.to_file(out, driver="GeoJSON")
    print(f"[1a] eGRID subregions: {len(gdf)} features → {out}")
    return len(gdf)


# ── 1b. eGRID emission rates + net generation ────────────────────────────────

def build_egrid_rates():
    # Emission rates from CSV
    rates_df = pd.read_csv(RAW / "egrid2023_rates.csv")
    acronym_col = "eGRID subregion acronym"
    co2_col = "Total output emission rates (lb/MWh)"

    rates_df[co2_col] = (
        rates_df[co2_col]
        .astype(str)
        .str.replace(",", "")
        .pipe(pd.to_numeric, errors="coerce")
    )

    # Net generation from eGRID2023_Data.xlsx Table 2
    wb = openpyxl.load_workbook(RAW / "eGRID2023_Data.xlsx", data_only=True, read_only=True)
    ws = wb["Table 2"]
    netgen = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        code = row[1]
        ng = row[4]
        if isinstance(code, str) and code not in ("", "eGRID subregion acronym", "U.S.") and isinstance(ng, (int, float)):
            netgen[code.strip()] = float(ng)
    wb.close()

    name_col = "eGRID subregion name"
    rates = {}
    for _, r in rates_df.iterrows():
        code = str(r[acronym_col]).strip()
        if not code or code in ("nan", "U.S."):
            continue
        co2 = r[co2_col]
        if pd.isna(co2):
            continue
        rates[code] = {
            "name":     str(r[name_col]).strip() if name_col in rates_df.columns else code,
            "co2_rate": float(co2),
            "net_gen":  netgen.get(code),
        }

    out = OUT / "egrid_rates.json"
    out.write_text(json.dumps(rates, indent=2))

    # National demand-weighted average
    total_gen = sum(v["net_gen"] for v in rates.values() if v["net_gen"])
    wavg = sum(v["co2_rate"] * v["net_gen"] for v in rates.values() if v["net_gen"]) / total_gen

    print(f"[1b] eGRID rates: {len(rates)} subregions → {out}")
    print(f"     National demand-weighted avg CO₂ rate: {wavg:.1f} lb/MWh")
    return wavg


# ── 1c. MECS electricity ─────────────────────────────────────────────────────

MECS_REGION_LABELS = ["Northeast", "Midwest", "South", "West"]

def build_mecs():
    wb = openpyxl.load_workbook(RAW / "mecs_electricity.xlsx", data_only=True, read_only=True)
    ws = wb["Table 7.6"]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    # Locate region section start rows
    region_starts = {}
    for i, row in enumerate(rows):
        for cell in row[:3]:
            if isinstance(cell, str):
                for label in MECS_REGION_LABELS:
                    if label in cell and label not in region_starts:
                        region_starts[label] = i

    # Sort regions by row order; each section ends where the next begins (or EOF)
    sorted_regions = sorted(region_starts.items(), key=lambda x: x[1])
    region_bounds = []
    for idx, (label, start) in enumerate(sorted_regions):
        end = sorted_regions[idx + 1][1] if idx + 1 < len(sorted_regions) else len(rows)
        region_bounds.append((label, start, end))

    target_prefixes = tuple(NAICS_SECTOR.keys())
    records = []

    for region, start, end in region_bounds:
        for row in rows[start:end]:
            raw_naics = str(row[0]).strip() if row[0] is not None else ""
            naics_clean = raw_naics.strip()
            if not any(naics_clean.startswith(p) for p in target_prefixes):
                continue
            sector = naics_sector(naics_clean)
            if sector is None:
                continue
            label_cell = str(row[1]).strip() if row[1] is not None else ""
            elec_raw = row[3]  # Electricity (million kWh) column
            if isinstance(elec_raw, (int, float)) and not (isinstance(elec_raw, float) and np.isnan(elec_raw)):
                elec_val = float(elec_raw)
            elif isinstance(elec_raw, str) and elec_raw.replace(",", "").replace(".", "").lstrip("-").isdigit():
                elec_val = float(elec_raw.replace(",", ""))
            else:
                elec_val = None  # suppressed: Q, W, *, D, --

            records.append({
                "naics": naics_clean,
                "sector_label": label_cell,
                "census_region": region,
                "elec_mkwh": elec_val,
            })

    out = OUT / "mecs_electricity.json"
    out.write_text(json.dumps(records, indent=2))

    by_region = {}
    for r in records:
        if r["elec_mkwh"] is not None:
            by_region[r["census_region"]] = by_region.get(r["census_region"], 0) + 1
    print(f"[1c] MECS rows: {len(records)} total")
    for region, count in sorted(by_region.items()):
        print(f"     {region}: {count} non-null rows")
    print(f"     → {out}")
    return records


# ── 1d. TRI facilities ───────────────────────────────────────────────────────

def normalize_col(name):
    """Strip leading number prefix like '30. ' from TRI column names."""
    import re
    return re.sub(r"^\d+\.\s*", "", str(name)).strip().upper()

def build_tri():
    df = pd.read_csv(RAW / "tri_facilities.csv", low_memory=False)
    df.columns = [normalize_col(c) for c in df.columns]

    # Find columns defensively
    def find_col(candidates):
        for c in candidates:
            if c in df.columns:
                return c
        raise KeyError(f"None of {candidates} found in TRI columns: {list(df.columns[:40])}")

    col_name   = find_col(["FACILITY NAME", "FACILITYNAME", "FACILITY"])
    col_lat    = find_col(["LATITUDE", "LAT"])
    col_lon    = find_col(["LONGITUDE", "LON", "LONG"])
    col_naics  = find_col(["PRIMARY NAICS", "PRIMARYNAICS", "NAICS"])
    col_state  = find_col(["ST", "STATE"])
    col_city   = find_col(["CITY"])

    print(f"[1d] TRI columns → name={col_name}, lat={col_lat}, lon={col_lon}, "
          f"naics={col_naics}, state={col_state}, city={col_city}")

    target_prefixes = (
        "3311", "3312", "3313", "3314", "32731",
        "3271", "3272", "32531", "32512", "33441",
    )

    df[col_naics] = df[col_naics].astype(str).str.strip()
    mask = df[col_naics].apply(lambda x: any(x.startswith(p) for p in target_prefixes))
    df = df[mask].copy()

    df[col_lat] = pd.to_numeric(df[col_lat], errors="coerce")
    df[col_lon] = pd.to_numeric(df[col_lon], errors="coerce")
    df = df.dropna(subset=[col_lat, col_lon])

    facilities = []
    seen = set()
    for _, r in df.iterrows():
        key = (str(r[col_name]).strip(), r[col_lat], r[col_lon])
        if key in seen:
            continue
        seen.add(key)
        sector = naics_sector(r[col_naics])
        if sector is None:
            continue
        facilities.append({
            "name":   str(r[col_name]).strip(),
            "lat":    float(r[col_lat]),
            "lon":    float(r[col_lon]),
            "naics":  str(r[col_naics]),
            "sector": sector,
            "state":  str(r[col_state]).strip(),
            "city":   str(r[col_city]).strip(),
        })

    # Spatial join: assign eGRID subregion to each facility
    shp = RAW / "egrid2023_subregions" / "eGRID2023_Subregions.shp"
    gdf_sub = gpd.read_file(shp)[["Subregion", "geometry"]]
    gdf_fac = gpd.GeoDataFrame(
        {"idx": range(len(facilities))},
        geometry=gpd.points_from_xy(
            [f["lon"] for f in facilities],
            [f["lat"] for f in facilities],
        ),
        crs="EPSG:4326",
    )
    joined = gpd.sjoin(gdf_fac, gdf_sub, how="left", predicate="within")
    joined = joined.drop_duplicates("idx")
    subr_by_idx = dict(zip(joined["idx"], joined["Subregion"]))
    for i, f in enumerate(facilities):
        subr = subr_by_idx.get(i)
        f["egrid_subregion"] = subr if isinstance(subr, str) else None

    out = OUT / "tri_facilities.json"
    out.write_text(json.dumps(facilities, indent=2))

    by_sector = {}
    for f in facilities:
        by_sector[f["sector"]] = by_sector.get(f["sector"], 0) + 1
    no_subr = sum(1 for f in facilities if not f["egrid_subregion"])
    print(f"[1d] TRI facilities: {len(facilities)} unique → {out}")
    for sector, count in sorted(by_sector.items()):
        print(f"     {sector}: {count}")
    print(f"     ({no_subr} facilities outside all subregion boundaries)")
    return facilities


# ── main ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=== FPFA build.py ===\n")
    n_subregions = build_geojson()
    wavg         = build_egrid_rates()
    mecs_rows    = build_mecs()
    facilities   = build_tri()

    print(f"\n=== Validation ===")
    print(f"  eGRID subregions:              {n_subregions}")
    print(f"  TRI facilities (unique):       {len(facilities)}")
    print(f"  MECS rows (all):               {len(mecs_rows)}")
    print(f"  National avg CO₂ rate:         {wavg:.1f} lb/MWh")
    print("\nDone. Serve with: python -m http.server")
